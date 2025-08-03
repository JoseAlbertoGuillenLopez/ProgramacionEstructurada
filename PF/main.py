import funciones
import getpass
from conexion import *
from usuarios import usuario
from clientes import cliente
from ventas import venta


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
import os

def main(): 
    usuarios_dicc = {"id": "", "nombre": "", "apellidos": "", "email":"", "password": ""}
    opcion = True
    while opcion:
        funciones.borrarPantalla()
        opcion = funciones.menu_principlal()
        match opcion:
            case "1" | "REGISTRO":
                funciones.borrarPantalla()
                print("\n \t ..::📝  Registro en el Sistema  📝::..")
                usuarios_dicc["nombre"] = input("\t ¿Cual es tu nombre?: ").upper().strip()
                usuarios_dicc["apellidos"] = input("\t ¿Cuales son tus apellidos?: ").upper().strip()
                usuarios_dicc["email"] = input("\t Ingresa tu email: ").lower().strip()
                usuarios_dicc["password"] = getpass.getpass("\t Ingresa tu contraseña: ").strip()
                registro = usuario.registrar(usuarios_dicc["nombre"], usuarios_dicc["apellidos"], usuarios_dicc["email"], usuarios_dicc["password"])
                if registro:
                    print(f"\n\t✅  {usuarios_dicc['nombre']} {usuarios_dicc['apellidos']}, se registro correctamente, con el email {usuarios_dicc['email']}  ✅")
                else:
                    print(f"\n\t...⚠️  Por favor intentelo de nuevo  ⚠️ ...")
                funciones.esperarTecla()
            case "2" | "LOGIN":
                funciones.borrarPantalla()
                print("\n \t ..::🔐  Inicio de Sesión  🔐::.. ")
                usuarios_dicc["email"] = input("\tIngresa tu E-mail: ").lower().strip()
                usuarios_dicc["password"] = getpass.getpass("\tIngresa tu contraseña: ").strip()
                login = usuario.login(usuarios_dicc["email"], usuarios_dicc["password"])
                if login:
                    opcion = input("\tDesea trabajar con clientes o ventas? (C/V): ").upper().strip()
                    if opcion == "C":
                        menu_clientes(login[0], login[1], login[2])
                    elif opcion == "V":
                        menu_ventas(login[0], login[1], login[2])
                    else:
                        print("⚠️  Opción no válida, vuelva a intentarlo.  ⚠️")
                else:
                    print("\n\t⚠️  Email y/o contrasena incorrecta, vuelva a intentarlo...  ⚠️")
                funciones.esperarTecla() 
            case "3" | "SALIR":
                opcion = False    
                funciones.borrarPantalla()
                input("\n\t\t 🥺​ Terminaste la ejecucion del SW 🥺​ ")
            case _: 
                input("\n\t\t 🚫 Opción invalida vuelva a intentarlo 🚫") 





def menu_clientes(usuario_id, nombre, apellidos):
    funciones.borrarPantalla()
    clientes_dicc = {"id":"", "id_usuario": usuario_id, "nombre": "", "tel": "", "direccion": "", "correo": "", "edad": ""}
    opcion = True
    while opcion:
        funciones.borrarPantalla()
        print(f"\n\t\t👋  Bienvenido {nombre} {apellidos}, has iniciado sesión ...")
        opcion = funciones.menu_clientes()
        match opcion:
            case "1" | "CREAR":
                funciones.borrarPantalla()
                print("\n \t .::✏️  Crear Cliente  ✏️::. ")
                clientes_dicc["nombre"] = input("\tNombre: ").upper().strip()
                while True:
                    clientes_dicc["tel"] = input("\tTeléfono: ").strip()
                    if clientes_dicc["tel"].isdigit():
                        break
                    else:
                        print("\t⚠️   Solo valores numericos  ⚠️ ")
                clientes_dicc["direccion"] = input("\tDirección: ").upper().strip()
                clientes_dicc["correo"] = input("\tCorreo: ").lower().strip()
                while True:
                    clientes_dicc["edad"] = input("\tEdad: ").strip()
                    if clientes_dicc["edad"].isdigit():
                        break
                    else:
                        print("\t⚠️   Solo valores numericos  ⚠️ ")
                registro = cliente.crear(clientes_dicc["id_usuario"], clientes_dicc["nombre"], clientes_dicc["tel"], clientes_dicc["direccion"], clientes_dicc["correo"], clientes_dicc["edad"])
                if registro:
                    print(f"\t✅  Cliente {clientes_dicc['nombre']} creado exitosamente.  ✅")
                else:
                    print("❌  Error al crear el cliente.  ❌")
                funciones.esperarTecla()
            case "2" | "CONSULTAR":
                funciones.borrarPantalla()
                print("\n \t .::📋  Consultar Cliente  📋::. ")
                id_cliente = input("\tIngrese el ID del cliente: ").strip()
                registro = cliente.consultar(id_cliente)
                if registro:
                    print(f"+{'-'*5}+{'-'*22}+{'-'*22}+{'-'*17}+{'-'*27}+{'-'*27}+{'-'*6}+")
                    print(f"| {'ID':<3} | {'Usuario ID':<20} | {'Nombre':<20} | {'Teléfono':<15} | {'Dirección':<25} | {'Correo':<25} | {'Edad':<4} |")
                    print(f"+{'-'*5}+{'-'*22}+{'-'*22}+{'-'*17}+{'-'*27}+{'-'*27}+{'-'*6}+")
                    print(f"| {registro[0]:<3} | {registro[1]:<20} | {registro[2]:<20} | {registro[3]:<15} | {registro[4]:<25} | {registro[5]:<25} | {registro[6]:<4} |")
                    print(f"+{'-'*5}+{'-'*22}+{'-'*22}+{'-'*17}+{'-'*27}+{'-'*27}+{'-'*6}+")
                else:
                    print("\t⚠️  Cliente no encontrado.  ⚠️") 
                funciones.esperarTecla()
            case "3" | "ACTUALIZAR":
                funciones.borrarPantalla()
                print("\n \t .::🔄  Actualizar Cliente  🔄::. ")
                registro = cliente.actualizar(usuario_id)
                funciones.esperarTecla()
            case "4" | "ELIMINAR":
                funciones.borrarPantalla()
                print("\n \t\t .::🗑️  Eliminar Cliente  🗑️::. ")
                registro = cliente.borrar()
                if registro:
                    print(f"✅  Cliente {registro} eliminado exitosamente.  ✅")
                funciones.esperarTecla()
            case "5" | "LISTAR":
                funciones.borrarPantalla()
                print("\n \t\t\t\t\t\t .::📋  Listar Clientes ::.  📋")
                registros = cliente.listar()
                if registros:
                    print(f"+{'-'*5}+{'-'*22}+{'-'*22}+{'-'*17}+{'-'*27}+{'-'*27}+{'-'*6}+")
                    print(f"| {'ID':<3} | {'Usuario ID':<20} | {'Nombre':<20} | {'Teléfono':<15} | {'Dirección':<25} | {'Correo':<25} | {'Edad':<4} |")
                    print(f"+{'-'*5}+{'-'*22}+{'-'*22}+{'-'*17}+{'-'*27}+{'-'*27}+{'-'*6}+")
                    for fila in registros:
                        print(f"| {fila[0]:<3} | {fila[1]:<20} | {fila[2]:<20} | {fila[3]:<15} | {fila[4]:<25} | {fila[5]:<25} | {fila[6]:<4} |")
                        print(f"+{'-'*5}+{'-'*22}+{'-'*22}+{'-'*17}+{'-'*27}+{'-'*27}+{'-'*6}+")
                funciones.esperarTecla()
            case "6" | "EXPORTAR":
                nombre_tabla = "clientes"  
                query = f"SELECT * FROM {nombre_tabla}"
                df = pd.read_sql(query, con=conexion)
                df.to_excel("clientes.xlsx", index=False, engine="openpyxl")
                os.startfile(os.getcwd())
            case "7" | "SALIR":
                opcion = False    
                funciones.borrarPantalla()
                print("\n\t\t 🥺 Volviendo al menu principal 🥺​ ")
            case _: 
                input("\n\t\t 🚫 Opción invalida vuelva a intentarlo 🚫")






def menu_ventas(usuario_id, nombre, apellidos): 
    funciones.borrarPantalla()
    ventas_dicc = {"id": "", "id_usuario": usuario_id, "id_cliente":"", "monto": "", "nprendas": "", "metodopago": "", "fecha": "" }
    opcion = True
    while opcion:
        funciones.borrarPantalla()
        print(f"\n \t \t👋  Bienvenido {nombre} {apellidos}, has iniciado sesión ...")
        opcion = funciones.menu_ventas()
        match opcion:
                case "1" | "CREAR":
                    funciones.borrarPantalla()
                    print("\n \t .::✏️  Crear Venta::.  ✏️")
                    ventas_dicc["id_cliente"] = input("\tID del cliente: ").upper().strip()
                    while True:
                        monto = input("\tMonto: ")
                        try:
                            monto_float = float(monto)
                            ventas_dicc["monto"] = monto_float
                            break
                        except ValueError:
                            print("\t⚠️  Solo valores numéricos  ⚠️")
                    while True:
                        ventas_dicc["nprendas"] = input("\tNumero de prendas: ")
                        if ventas_dicc["nprendas"].isdigit():
                            break
                        else:
                            print("\t⚠️  Solo valores numericos  ⚠️")
                    while True:
                        ventas_dicc["metodopago"] = input("\tMetodo de pago (Efectivo=E/Transfrenecia=T): ").lower().strip()
                        if ventas_dicc["metodopago"]=="e":
                            ventas_dicc["metodopago"]="EFECTIVO"
                            break
                        elif ventas_dicc["metodopago"]=="t":
                            ventas_dicc["metodopago"]="TRANSFERENCIA"
                            break
                        else:
                            print("\t⚠️  Solo Efectivo=E/Transfrenecia=T  ⚠️")
                    registro = venta.crear(usuario_id,ventas_dicc["id_cliente"],ventas_dicc["monto"],ventas_dicc["nprendas"],ventas_dicc["metodopago"])
                    if registro:
                        print(f"\t✅  Venta creada exitosamente.  ✅")
                    else:
                        print("❌  Error al crear la venta.  ❌")
                    funciones.esperarTecla()
                case "2" | "CONSULTAR":
                    funciones.borrarPantalla()
                    print("\n \t .::📋  Consultar Venta  📋::. ")
                    id_venta= input("\tIngrese el ID del Venta: ").strip()
                    registro = venta.consultar(id_venta)
                    if registro:
                        print(f"+{'-'*5}+{'-'*12}+{'-'*13}+{'-'*10}+{'-'*15}+{'-'*20}+{'-'*12}+")
                        print(f"| {'ID':<3} | {'ID Usuario':<10} | {'ID Cliente':<11} | {'Monto':<8} | {'# Prendas':<13} | {'Método de Pago':<18} | {'Fecha':<10} |")
                        print(f"+{'-'*5}+{'-'*12}+{'-'*13}+{'-'*10}+{'-'*15}+{'-'*20}+{'-'*12}+")
                        print(f"| {registro[0]:<3} | {registro[1]:<10} | {registro[2]:<11} | {registro[3]:<8.2f} | {registro[4]:<13} | {registro[5]:<18} | {str(registro[6]):<10} |")
                        print(f"+{'-'*5}+{'-'*12}+{'-'*13}+{'-'*10}+{'-'*15}+{'-'*20}+{'-'*12}+")
                    else:
                        print("\t⚠️  Venta no encontrada.  ⚠️")
                    funciones.esperarTecla()
                case "3" | "ACTUALIZAR":
                    funciones.borrarPantalla()
                    print("\n \t .::🔄  Actualizar Ventas  🔄::. ")
                    registro = venta.actualizar(usuario_id)
                    if registro:
                        print(f"✅  Venta {registro} actualizada exitosamente.  ✅")
                    funciones.esperarTecla()
                case "4" | "ELIMINAR":
                    funciones.borrarPantalla()
                    print("\n \t .::🗑️ Eliminar Venta  🗑️::. ")
                    registro = venta.borrar()
                    if registro:
                        print(f"✅  Venta {registro} eliminada exitosamente.  ✅")
                    funciones.esperarTecla()
                case "5" | "LISTAR":
                    funciones.borrarPantalla()
                    print("\n \t .:: 📋  Listar Ventas   📋::. ")
                    registros = venta.listar()
                    if registros:
                        print(f"+{'-'*5}+{'-'*12}+{'-'*13}+{'-'*10}+{'-'*15}+{'-'*20}+{'-'*12}+")
                        print(f"| {'ID':<3} | {'ID Usuario':<10} | {'ID Cliente':<11} | {'Monto':<8} | {'# Prendas':<13} | {'Método de Pago':<18} | {'Fecha':<10} |")
                        print(f"+{'-'*5}+{'-'*12}+{'-'*13}+{'-'*10}+{'-'*15}+{'-'*20}+{'-'*12}+")
                        for i in registros:
                            print(f"| {i[0]:<3} | {i[1]:<10} | {i[2]:<11} | {i[3]:<8.2f} | {i[4]:<13} | {i[5]:<18} | {str(i[6]):<10} |")
                            print(f"+{'-'*5}+{'-'*12}+{'-'*13}+{'-'*10}+{'-'*15}+{'-'*20}+{'-'*12}+")
                    funciones.esperarTecla()
                case "6" | "EXPORTAR":
                    df = pd.read_sql("SELECT * FROM ventas", con=conexion)
                    df["fecha"] = pd.to_datetime(df["fecha"]).dt.date
                    archivo = "ventas.xlsx"
                    df.to_excel(archivo, index=False, engine="openpyxl")
                    wb = load_workbook(archivo)
                    ws = wb.active
                    for col in ws.columns:
                        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max_length + 2
                    wb.save(archivo)
                    os.startfile(os.getcwd())
                case "7" | "SALIR":
                    opcion = False    
                    funciones.borrarPantalla()
                    print("\n\t\t 🥺 Volviendo al menu principal 🥺​ ") 
                case _: 
                    input("\n\t\t 🚫 Opción invalida vuelva a intentarlo 🚫")



if __name__=="__main__":
    main()
    